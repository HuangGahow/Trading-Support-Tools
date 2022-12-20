# -*- coding: utf-8 -*-
"""
Created on Sat Nov  5 22:05:48 2022

@author: HP
"""
# In[1]


import pandas as pd
import numpy as np
import os


import openpyxl as op
np.set_printoptions(suppress=True)

templatepath = input("please input your template's path:")
path = input("please input your 北向国债期货连接端's path:")

#templatepath = r"D:\Users\Huangjiahao\Desktop\FIOP\Template.xlsx"

#path = r"D:\Users\Huangjiahao\Desktop\FIOP\北向国债期货连接端20221205.xlsx"

folder = os.getcwd() + '\\FPsheet\\'
if not os.path.exists(folder):
    os.makedirs(folder)

uwfuture = pd.read_excel(path, sheet_name="国债期货结算")
uwfuture = uwfuture.dropna(axis=0, subset=['交易编号'])
Date = uwfuture['平仓日期'].iloc[-1]
uwfuture = uwfuture.loc[uwfuture['平仓日期'] == Date]
Bflag = "Notional Amount x (Final Price- Initial Price)/100"
Sflag = "Notional Amount x (Initial Price- Final Price)/100"

uwfuture['剩余名义本金（不含价格）']=uwfuture['剩余名义本金（不含价格）'].apply(lambda x: round(x,1)).fillna(0)
uwfuture['结算名义本金（不含价格）']=uwfuture['结算名义本金（不含价格）'].apply(lambda x: round(x,1)).fillna(0)
uwfuture['结算金额']=uwfuture['结算金额'].apply(lambda x: round(x,1))
uwfuture.drop(['Unnamed: 0','Unnamed: 3','Unnamed: 10','Unnamed: 11'],axis = 1, inplace = True)
uwfuture['Contract'] = uwfuture['交易编号'].apply(lambda x: x.split('-')[-3])
uwfuture['FPornot'] =  uwfuture['交易编号'].apply(lambda x: "FP" in x)



FP = uwfuture[uwfuture['FPornot']==True].reset_index() 
FP['剩余名义本金（不含价格）'].fillna(0)
FP['B/S'] = FP['交易编号'].apply(lambda x: x.split('-')[-1].lower())
FP['B/Sflag'] = FP['B/S'].apply(lambda x: Bflag if x=="buy" else Sflag)
FP['PUNA'] = FP['结算名义本金（不含价格）']/1000000
FP['ONApost'] = FP['剩余名义本金（不含价格）']/1000000
FP['ONAprior'] = FP['PUNA'] + FP['ONApost']
FP['PUNA'] = FP['PUNA'].apply(lambda x: str(int(x))+'mio')
FP['ONApost'] = FP['ONApost'].apply(lambda x: str(int(x))+'mio')
FP['ONAprior'] = FP['ONAprior'].apply(lambda x: str(int(x))+'mio')

FPlist = FP['交易编号']

if len(FPlist)==0:
    print("NO FP")
    
else:
    wb = op.load_workbook(templatepath)
    for i in range(len(FPlist)):
        openbond = wb['Unwind Detail']
        openbond.cell(5, 2 , FP['平仓日期'].iloc[i])
        openbond.cell(7, 2 , FP['交易编号'].iloc[i])
        openbond.cell(8, 2 , FP['Contract'].iloc[i])
        openbond.cell(9, 2 , FP['ONAprior'].iloc[i])
        openbond.cell(10, 2 , FP['PUNA'].iloc[i])
        openbond.cell(11, 2 , FP['ONApost'].iloc[i])
        openbond.cell(12, 2 , FP['Unnamed: 17'].iloc[i])
        openbond.cell(13, 2 , FP['Unnamed: 4'].iloc[i])
        openbond.cell(14, 2 , FP['平仓日期'].iloc[i])
        openbond.cell(15, 2 , FP['Unnamed: 5'].iloc[i])
        openbond.cell(16, 2 , FP['B/Sflag'].iloc[i])
        openbond.cell(17, 2 , FP['结算金额'].iloc[i])
        
        img = op.drawing.image.Image('CICCLOGO.jpg')
        img.anchor = 'A1'
        img.width = 190
        img.height = 65
        openbond.add_image(img)
        
        #wb.save('D:/Users/Huangjiahao/Desktop/FIOP/FPsheet/'+ FPlist[i] + '.xlsx')
        wb.save(folder+ 'FP-中金公司-Unwind+'+ FPlist[i] + '.xlsx')
        print("Done")

print("FPsheets completed")