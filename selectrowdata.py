# -*- coding: utf-8 -*-
"""
Created on Fri Nov 18 09:03:26 2022

@author: Huangjiahao
"""

import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl as op

path1 =input("please input your 交易合约明细's path:")
path2 =input("please input your 北向国债期货连接端's path:")

rawdata = pd.read_excel(io=path1,header = 0 )
todaydate = rawdata.iloc[0,0]
rawdata = rawdata[rawdata['交易日期']==todaydate]
bond = rawdata[rawdata['开关仓']=='open']
unwind = rawdata[rawdata['开关仓']=='unwind']

#change time into dd/mm/yyyy
def changetime(x):
    x = str(x)
    format = '%Y%m%d'
    datetime1 = datetime.strptime(x[:-6], format)
    #datetime1 = datetime.strftime('%d/%m/%Y',datetime1)
    return datetime1
todaydate = changetime(todaydate)

def valdate(x):
    if x == '2303':
        date = datetime.strptime('2/22/2023','%m/%d/%Y')
    else:
        date =  datetime.strptime('11/15/2022','%m/%d/%Y')        
    return date

def cleandata(bond):
    bond = bond.loc[:,['合约编号','BSFLAG','STKID','数量','交易价格grossprice','对手方']]
    bond['date1'] = bond['STKID'].apply(lambda x: x[-4:])
    bond['valuationdate']=bond['date1'].apply(lambda x: valdate(x))
    bond.loc[(bond['date1']=='2212') & ((bond['对手方']=='GSI')|(bond['对手方']=='MS')), 'valuationdate'] \
        =datetime.strptime('11/25/2022','%m/%d/%Y')

    return bond

bond = cleandata(bond)
unwind = cleandata(unwind)


class rawdata1():
    def __init__(self,bond):
        self.bond = bond
        self.date1 = self.bond['STKID'].apply(lambda x: x[-4:])
        self.valuationdate = self.bond['valuationdate'].reset_index(drop=True)
        self.referencenumber = self.bond['合约编号'].apply(lambda x:'xb-'+x).reset_index(drop=True)
        self.BSFLAG = self.bond['BSFLAG'].reset_index(drop=True)
        self.STKID = self.bond['STKID'].reset_index(drop=True)
        self.Number = self.bond['数量'].apply(lambda x:float(x)).reset_index(drop=True)
        self.p0 = self.bond['交易价格grossprice'].apply(lambda x:float(x)).reset_index(drop=True)
        
bond = rawdata1(bond)
unwind = rawdata1(unwind)

# get length of bond




#write into excel 
wb = op.load_workbook(path2)

openbond = wb['Bond']
bond1 = pd.DataFrame(openbond.values)
bond1.dropna(axis=0,subset=[3],inplace=True)
bondmaxrow = len(bond1)


for i in range(1,len(bond.referencenumber)+1):
    openbond.cell(bondmaxrow + i, 4 ,bond.referencenumber[i-1])
for i in range(1,len(bond.BSFLAG)+1):
    openbond.cell(bondmaxrow + i, 5 ,bond.BSFLAG[i-1])
for i in range(1,len(bond.STKID)+1):
    openbond.cell(bondmaxrow + i, 7 ,bond.STKID[i-1])
for i in range(1,len(bond.Number)+1):
    openbond.cell(bondmaxrow + i, 8 ,bond.Number[i-1])
for i in range(1,len(bond.p0)+1):
    openbond.cell(bondmaxrow + i, 9 ,bond.p0[i-1])
for i in range(1,len(bond.referencenumber)+1):
    openbond.cell(bondmaxrow + i, 13 ,todaydate)
for i in range(1,len(bond.valuationdate)+1):
    openbond.cell(bondmaxrow + i, 14 ,bond.valuationdate[i-1])




openbondx = wb['国债期货结算']
unwind1 = pd.DataFrame(openbondx.values)
unwind1.dropna(axis=0,subset=[2],inplace=True)
unwindmaxrow = len(unwind1)
for i in range(1,len(unwind.referencenumber)+1):
    openbondx.cell(unwindmaxrow + i, 3 ,unwind.referencenumber[i-1])
for i in range(1,len(unwind.p0)+1):
    openbondx.cell(unwindmaxrow+ i, 6 ,unwind.p0[i-1])
for i in range(1,len(unwind.referencenumber)+1):
    openbondx.cell(unwindmaxrow + i, 2 ,todaydate)
for i in range(1,len(unwind.Number)+1):
    openbondx.cell(unwindmaxrow + i, 10 ,unwind.Number[i-1])





wb.save('New.xlsx')
print("Done")
